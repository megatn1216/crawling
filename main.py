from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from PyQt5 import QtGui
import sys
from selenium import webdriver
import openpyxl
import datetime as dt
import time
from bs4 import BeautifulSoup
import requests

# import pandas as pd
import re
from openpyxl.styles import PatternFill
# import io
 
class Window(QWidget):
    def __init__(self):
        super().__init__()
 
        # label text
        self.label = QLabel("Hello World!")
        self.label.setFont(QtGui.QFont('Hack', 13))
        
        # 텍스트 입력 (brand명)
        self.lineEdit = QLineEdit()
        self.lineEdit.setMaximumHeight(100)
        self.lineEdit.setPlaceholderText("enter brand")

        # button
        btn1 = QPushButton('시작', self)
        btn1.setStyleSheet('background:lightgrey;')
        btn1.setMaximumHeight(100)
        btn1.clicked.connect(self.btn1_clicked)
        
        btn2 = QPushButton('시작(사진포함)', self)
        btn2.setStyleSheet('background:lightgrey;')
        btn2.setMaximumHeight(100)
        btn2.clicked.connect(self.btn2_clicked)

        btnC = QPushButton('닫기', self)
        btnC.setStyleSheet('background:lightgrey;')
        btnC.setMaximumHeight(100)
        btnC.clicked.connect(self.btn_win_close)

        # layout
        vbox = QVBoxLayout()

        vbox.addWidget(self.label, alignment=Qt.AlignCenter)
        vbox.addWidget(self.lineEdit, alignment=Qt.AlignCenter)
        vbox.addWidget(btn1)
        vbox.addWidget(btn2)
        vbox.addWidget(btnC)

        self.setLayout(vbox)


        # window 셋팅
        self.setWindowFlags(Qt.WindowStaysOnTopHint|Qt.FramelessWindowHint)
        self.setGeometry(-1100, 500, 300, 200)
        # self.setWindowTitle("Crawling_oord")

    def btn_win_close(self):
        self.close()

    def btn1_clicked(self):
        # 기존 파일 로딩
        fname = QFileDialog.getOpenFileName(self)
        self.label.setText(fname[0])
        data = openpyxl.load_workbook(fname[0], data_only=True)


        # ##### 임시코드
        # temp_path ='C:/Users/jagua/Desktop/Crawling_oor/crawling/crawling_1116.xlsx'
        # data = openpyxl.load_workbook(temp_path, data_only=True)
        # #############
        
        # sheet 이름 전체
        sheetNames = data.sheetnames
        

        # ####### Selenium 
        chrome_options = webdriver.ChromeOptions()

        # 내부창을 띄우지 않게
        chrome_options.add_argument('--headless')

        # chrome driver 생성
        driver = webdriver.Chrome('chromedriver',chrome_options=chrome_options)

        # open 대기 3초
        # driver.implicitly_wait(3)
        # #########
        
        # brandList = ['all-global', 'kasumi', 'kyocera-ceramic-knives', 'all-le-creuset', 'all-mauviel', 'messermeister_store.htm','all-miyabi', 'murray-carter-by-spyderco', 'opinel', 'scanpan', 'all-shun', 'all-staub', 'tojiro', 'forschner_knives.htm', 'viking', 'all-wusthof', 'all-yaxell', 'all-zwilling-ja-henckels']
        brandList = self.lineEdit.text().split(',')

        for brand in brandList:
            strtime = dt.datetime.now().strftime('%m%d')

            url = f"https://www.cutleryandmore.com/" + brand
            # selenium으로 url 접속
            driver.get(url)
            
            # beautifulsoup은 parsing만
            soup = BeautifulSoup(driver.page_source, 'html.parser')

            # 파싱 결과 중 info class로 select
            frame = soup.select('.info') 

            # 해당 sheet가 있는지 확인
            # 없으면 시트 생성
            if brand not in sheetNames :
                data.create_sheet(brand)

                # 상품명 column 추가                
                data[brand].cell(1, 1).value = 'brand'
                data[brand].cell(1, 2).value = 'upload'
                data[brand].cell(1, 3).value = 'name'
                print('[sheet 생성] : ', brand)
                
            # 기존 data 최대값
            maxR = data[brand].max_row
            maxC = data[brand].max_column
            
            # 전체 아이템 개수 세기
            num=0
            
            # 오늘자 price, stock column 추가
            data[brand].cell(1, maxC+1).value = strtime+'_pr'
            data[brand].cell(1, maxC+2).value = strtime+'_st'


            for i in frame :
                # 상품 명
                name = i.find('span', attrs={'class':'name'}).text
                
                # 상품 가격
                if (i.find('span', attrs={'class':'price'})) :
                    pre_price = i.find('span', attrs={'class':'price'}).text
                    regex = re.compile("{}(.*){}".format(re.escape('\t\t\t\t\t$'), re.escape('\t\t\t\t\t')));
                    # 전처리 이후 가격
                    aft_price = regex.findall(pre_price)[0]
                else :
                    # 가격이 없을때 default
                    aft_price = 0

                # 재고 수 
                if (i.find('span', attrs={'class':'stockLevel'})) :
                    stock = i.find('span', attrs={'class':'stockLevel'}).text
                else :
                    stock = 0

                # print('[Item 정보] ', name, '(price)', aft_price, '(stock)', stock)


                flag=0
                for j in range(2, maxR+1):
                    # Item이 있을 경우 값 추가       
                    if (name == data[brand].cell(j,3).value):
                        data[brand].cell(j, maxC+1).value = aft_price 
                        data[brand].cell(j, maxC+2).value = stock
                        flag=1
                        nextNum = float(data[brand].cell(j, maxC+1).value)
                        prevNum = float(data[brand].cell(j, maxC-1).value)
                        if (prevNum != nextNum):
                            print("[Item 존재! 가격 다름!10] ", name, "//", prevNum, "//", nextNum)
                            if (nextNum > prevNum): #커졌으면 빨강
                                data[brand].cell(j, maxC+1).fill = PatternFill(start_color='ff9778', end_color='ff9778', fill_type='solid')
                            else: #작아졌으면 파랑
                                data[brand].cell(j, maxC+1).fill = PatternFill(start_color='8cabff', end_color='8cabff', fill_type='solid')

                # Item이 없을 경우 Row 추가
                if (flag == 0) : 
                    last_row = data[brand].max_row
                    print("[Item 없음! Row 추가!] ", name)
                    # 없을경우 열추가 해야함
                    data[brand].cell(last_row+1, 1).value = brand
                    data[brand].cell(last_row+1, 3).value = name
                    data[brand].cell(last_row+1, maxC+1).value = aft_price
                    data[brand].cell(last_row+1, maxC+2).value = stock

                data[brand].column_dimensions['C'].width = 50
                num += 1 # item 수 세기
                
            print("[" + brand + " : complete(" + str(num) + ")]")
            self.label.setText("[" + brand + " : complete(" + str(num) + ")]")

        data.save(f'crawling_{strtime}.xlsx')

    def btn2_clicked(self):
        # 기존 파일 로딩
        fname = QFileDialog.getOpenFileName(self)
        self.label.setText(fname[0])
        data = openpyxl.load_workbook(fname[0], data_only=True)


        # ##### 임시코드
        # temp_path ='C:/Users/jagua/Desktop/Crawling_oor/crawling/crawling_1116.xlsx'
        # data = openpyxl.load_workbook(temp_path, data_only=True)
        # #############
        
        # sheet 이름 전체
        sheetNames = data.sheetnames
        

        # ####### Selenium 
        chrome_options = webdriver.ChromeOptions()

        # 내부창을 띄우지 않게
        chrome_options.add_argument('--headless')

        # chrome driver 생성
        driver = webdriver.Chrome('chromedriver',chrome_options=chrome_options)

        # open 대기 3초
        # driver.implicitly_wait(3)
        # #########
        
        # brandList = ['all-global', 'kasumi', 'kyocera-ceramic-knives', 'all-le-creuset', 'all-mauviel', 'messermeister_store.htm','all-miyabi', 'murray-carter-by-spyderco', 'opinel', 'scanpan', 'all-shun', 'all-staub', 'tojiro', 'forschner_knives.htm', 'viking', 'all-wusthof', 'all-yaxell', 'all-zwilling-ja-henckels']
        brandList = self.lineEdit.text().split(',')

        for brand in brandList:
            strtime = dt.datetime.now().strftime('%m%d')

            url = f"https://www.cutleryandmore.com/" + brand
            # selenium으로 url 접속
            driver.get(url)
            
            # beautifulsoup은 parsing만
            soup = BeautifulSoup(driver.page_source, 'html.parser')

            # 파싱 결과 중 info class로 select
            frame = soup.select('.info') 

            # 해당 sheet가 있는지 확인
            # 없으면 시트 생성
            if brand not in sheetNames :
                data.create_sheet(brand)

                # 상품명 column 추가                
                data[brand].cell(1, 1).value = 'brand'
                data[brand].cell(1, 2).value = 'upload'
                data[brand].cell(1, 3).value = 'name'
                print('[sheet 생성] : ', brand)
                
            # 기존 data 최대값
            maxR = data[brand].max_row
            maxC = data[brand].max_column
            
            # 전체 아이템 개수 세기
            num=0
            
            # 오늘자 price, stock column 추가
            data[brand].cell(1, maxC+1).value = strtime+'_pr'
            data[brand].cell(1, maxC+2).value = strtime+'_st'


            for i in frame :
                # 상품 명
                name = i.find('span', attrs={'class':'name'}).text
                
                # 상품 가격
                if (i.find('span', attrs={'class':'price'})) :
                    pre_price = i.find('span', attrs={'class':'price'}).text
                    regex = re.compile("{}(.*){}".format(re.escape('\t\t\t\t\t$'), re.escape('\t\t\t\t\t')));
                    # 전처리 이후 가격
                    aft_price = regex.findall(pre_price)[0]
                else :
                    # 가격이 없을때 default
                    aft_price = 0

                # 재고 수 
                if (i.find('span', attrs={'class':'stockLevel'})) :
                    stock = i.find('span', attrs={'class':'stockLevel'}).text
                else :
                    stock = 0

                # print('[Item 정보] ', name, '(price)', aft_price, '(stock)', stock)


                flag=0
                for j in range(2, maxR+1):
                    # Item이 있을 경우 값 추가       
                    if (name == data[brand].cell(j,3).value):
                        data[brand].cell(j, maxC+1).value = aft_price 
                        data[brand].cell(j, maxC+2).value = stock
                        flag=1
                        nextNum = float(data[brand].cell(j, maxC+1).value)
                        prevNum = float(data[brand].cell(j, maxC-1).value)
                        if (prevNum != nextNum):
                            print("[Item 존재! 가격 다름!10] ", name, "//", prevNum, "//", nextNum)
                            if (nextNum > prevNum): #커졌으면 빨강
                                data[brand].cell(j, maxC+1).fill = PatternFill(start_color='ff9778', end_color='ff9778', fill_type='solid')
                            else: #작아졌으면 파랑
                                data[brand].cell(j, maxC+1).fill = PatternFill(start_color='8cabff', end_color='8cabff', fill_type='solid')

                # Item이 없을 경우 Row 추가
                if (flag == 0) : 
                    last_row = data[brand].max_row
                    print("[Item 없음! Row 추가!] ", name)
                    # 없을경우 열추가 해야함
                    data[brand].cell(last_row+1, 1).value = brand
                    data[brand].cell(last_row+1, 3).value = name
                    data[brand].cell(last_row+1, maxC+1).value = aft_price
                    data[brand].cell(last_row+1, maxC+2).value = stock

                data[brand].column_dimensions['C'].width = 50
                num += 1 # item 수 세기
                
            print("[" + brand + " : complete(" + str(num) + ")]")
            self.label.setText("[" + brand + " : complete(" + str(num) + ")]")

        data.save(f'crawling_{strtime}.xlsx')

app = QApplication(sys.argv)
root = Window()
root.show()
 
sys.exit(app.exec_())